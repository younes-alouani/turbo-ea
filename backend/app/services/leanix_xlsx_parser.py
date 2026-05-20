"""LeanIX xlsx (workspace export) parser.

LeanIX customers most often export their workspace as a multi-sheet
Excel workbook (the "Reports → Full Export" feature). The shape is
completely different from the tenant-cloning JSON snapshot handled in
:mod:`leanix_snapshot_parser`:

- One sheet per fact-sheet type (``Application``, ``BusinessCapability``,
  ``Process``, …). Column names are stable LeanIX field keys.
- One sheet per relation type (``childParentRelation``,
  ``applicationITComponentRelation``, plus tenant-defined relations
  whose sheet names get truncated to 31 chars by Excel — use the
  ``type`` column inside the sheet for the canonical name).
- Per-FS-row tag columns formatted as ``tags:<GroupName>`` with
  comma-separated values, subscription columns formatted as
  ``subscriptions:<RoleType>:<RoleName>`` with comma-separated emails.
- Auxiliary sheets: ``TagGroups``, ``Tags``, ``Documents``,
  ``Comments``, ``Types`` (enum option lists), ``ReadMe`` (skipped).

This parser is **schema-tolerant**: anything outside the documented
LeanIX core columns flows through to :attr:`FactSheet.custom_fields`
unchanged. Tenant-defined fact-sheet types, tag groups, and relation
types all surface via the existing metamodel-staging pipeline so the
admin can map them post-import.

Relation rows reference endpoints by ``(displayName, factSheetType)``
rather than by id. The parser builds a per-type display-name lookup
during the first pass and resolves endpoints in a second pass.
Ambiguous display names (very rare in practice) pick the first match
and log a parse error so the admin sees it in the migration summary.
"""

from __future__ import annotations

import hashlib
import logging
from datetime import date, datetime
from io import BytesIO
from typing import Any, BinaryIO

from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.services.leanix_snapshot_parser import (
    Comment,
    Document,
    FactSheet,
    LeanixSnapshot,
    MetamodelField,
    MetamodelRelationType,
    MetamodelType,
    Relation,
    Subscription,
    Tag,
    UserRef,
)

logger = logging.getLogger(__name__)

# Fact-sheet column keys that map 1:1 to ``FactSheet`` dataclass slots.
# Anything outside this set (after stripping the dynamic ``lifecycle:``,
# ``tags:`` and ``subscriptions:`` prefixes) is treated as a custom
# attribute.
_FS_CORE_COLS: frozenset[str] = frozenset(
    {
        "id",
        "type",
        "name",
        "displayName",
        "status",
        "description",
        "category",
        "completion",
        "qualitySeal",
        "lxState",
        "createdAt",
        "updatedAt",
        # ``naFields`` is LeanIX's "intentionally not set" marker — we
        # ignore it on import; the absence of a value on the importing
        # side is itself the "not set" signal.
        "naFields",
        # ``level`` / ``childLevel`` are computed in LeanIX from the
        # hierarchy; Turbo EA derives its own, so don't import.
        "level",
        "childLevel",
        # ACL columns — Turbo EA's permission model is wholly different.
        "permittedReadACL",
        "permittedWriteACL",
    }
)

# LeanIX export sheet names that are not fact-sheet data.
_AUX_SHEETS: frozenset[str] = frozenset(
    {"ReadMe", "TagGroups", "Tags", "Documents", "Comments", "Types"}
)

# Relation sheets always include these columns. Anything else on the
# relation row becomes attributes on the resulting Relation.
_REL_CORE_COLS: frozenset[str] = frozenset(
    {
        "id",
        "type",
        "fromRelatedFactSheetDisplayName",
        "fromRelatedFactSheetType",
        "toRelatedFactSheetDisplayName",
        "toRelatedFactSheetType",
        "status",
        "activeFrom",
        "activeUntil",
        "naFields",
    }
)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def parse_xlsx_path(path: str) -> LeanixSnapshot:
    """Parse a LeanIX workbook export from disk.

    Always reads via :class:`BytesIO` rather than handing openpyxl the
    path directly — openpyxl validates the extension and rejects files
    that don't end in ``.xlsx`` / ``.xlsm`` / etc. Migration uploads
    are stored under their migration id with a ``.bin`` suffix, so
    content-based loading is the only path that works for both
    ``upload.xlsx`` and stored snapshots.
    """
    with open(path, "rb") as fh:
        raw = fh.read()
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    return _parse_workbook(wb)


def parse_xlsx(stream: BinaryIO) -> LeanixSnapshot:
    """Parse a LeanIX workbook export from a binary stream."""
    raw = stream.read()
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    return _parse_workbook(wb)


def is_xlsx_payload(prefix: bytes) -> bool:
    """Return ``True`` if the byte prefix looks like an xlsx workbook.

    xlsx files are ZIP archives — magic ``PK\\x03\\x04``. We only need
    the first four bytes.
    """
    return len(prefix) >= 4 and prefix[:4] == b"PK\x03\x04"


# ---------------------------------------------------------------------------
# Workbook walker
# ---------------------------------------------------------------------------


def _parse_workbook(wb: Any) -> LeanixSnapshot:
    errors: list[str] = []

    fact_sheets: list[FactSheet] = []
    relations: list[Relation] = []
    subscriptions: list[Subscription] = []
    documents: list[Document] = []
    comments: list[Comment] = []

    # Build the (display_name, fs_type) → leanix_id lookup as we walk
    # the fact-sheet sheets; relation/document/comment endpoints resolve
    # against it in a second pass.
    display_index: dict[tuple[str, str], str] = {}
    seen_columns_per_type: dict[str, dict[str, list[Any]]] = {}

    # Tag-group + tag dicts go through their own sheets but FS rows
    # reference tags by ``<groupName>:<tagName>`` — build the inverse
    # lookup so we can populate ``FactSheet.tags`` with stable ids.
    tag_groups: dict[str, dict[str, Any]] = {}
    tag_records: list[Tag] = []
    tag_by_group_and_name: dict[tuple[str, str], str] = {}

    # First pass: enum option lists from the ``Types`` sheet (one row
    # per ``(factSheetType, fieldName)`` with the enum values spread
    # across the trailing columns). LeanIX only populates this sheet
    # for fields that have at least one card actually using the field
    # — it can underrepresent the real enum.
    types_options: dict[tuple[str, str], list[str]] = {}
    if "Types" in wb.sheetnames:
        for row in _data_rows(wb["Types"], skip_label_row=True):
            fs_type = _str_or_none(row.get("factSheetType"))
            field_name = _str_or_none(row.get("fieldName"))
            if not fs_type or not field_name:
                continue
            values: list[str] = []
            for col, val in row.items():
                if col in {"factSheetType", "fieldName"}:
                    continue
                s = _str_or_none(val)
                if s:
                    values.append(s)
            if values:
                types_options[(fs_type, field_name)] = values

    # Second pass: the ``ReadMe`` sheet — the authoritative LeanIX
    # field reference. Carries the **complete** enum list for every
    # field (not just what the data happens to use), the data type
    # (String / Integer / Percent / Datetime / Boolean / String list)
    # and the mandatory flag, organised by fact-sheet type sections.
    # Used as the primary source for field metadata when synthesising
    # the metamodel, with ``Types`` as fallback.
    readme_fields = _parse_readme(wb)

    # Tag groups + tags (both metadata sheets).
    if "TagGroups" in wb.sheetnames:
        for row in _data_rows(wb["TagGroups"], skip_label_row=True):
            name = _str_or_none(row.get("name"))
            if not name:
                continue
            tag_groups[name] = {
                "leanix_id": _str_or_none(row.get("id")) or name,
                "name": name,
                "mode": _str_or_none(row.get("mode")) or "MULTIPLE",
                "restrict_to_types": _str_or_none(row.get("restrictToFactSheetTypes")),
            }

    if "Tags" in wb.sheetnames:
        for row in _data_rows(wb["Tags"], skip_label_row=True):
            tag_id = _str_or_none(row.get("id"))
            name = _str_or_none(row.get("name"))
            # ``tagGroupId`` in the export holds the *group name*, not
            # the group uuid. Don't be surprised by that.
            group_name = _str_or_none(row.get("tagGroupId"))
            if not (tag_id and name and group_name):
                continue
            group_info = tag_groups.get(group_name) or {}
            tag_records.append(
                Tag(
                    leanix_id=tag_id,
                    name=name,
                    group_name=group_name,
                    group_mode=group_info.get("mode") or "MULTIPLE",
                    color=_str_or_none(row.get("backgroundColor")),
                )
            )
            tag_by_group_and_name[(group_name, name)] = tag_id

    # Second pass: each fact-sheet sheet.
    for sheet_name in wb.sheetnames:
        if sheet_name in _AUX_SHEETS:
            continue
        sheet = wb[sheet_name]
        first_header = _first_data_header(sheet)
        if first_header is None:
            continue
        # Heuristic: relation sheets always carry a ``type`` column AND
        # the ``fromRelatedFactSheetDisplayName`` column. Everything
        # else is a fact-sheet sheet.
        if "fromRelatedFactSheetDisplayName" in first_header:
            continue  # handled in the relations pass below

        for row in _data_rows(sheet, skip_label_row=True):
            try:
                fs = _build_fact_sheet(row, errors)
            except _SkipRowError:
                continue
            fact_sheets.append(fs)

            display_index[(_norm_name(fs.display_name or fs.name), fs.type)] = fs.leanix_id

            # Track every column seen per type for synthetic-metamodel emission.
            cols = seen_columns_per_type.setdefault(fs.type, {})
            for k in row.keys():
                if k not in cols:
                    cols[k] = []

            # Resolve tag columns into leanix tag ids.
            for col, value in row.items():
                if not (col and col.startswith("tags:")):
                    continue
                group_name = col[len("tags:") :]
                for tag_name in _split_csv(value):
                    tag_id = tag_by_group_and_name.get((group_name, tag_name))
                    if tag_id and tag_id not in fs.tags:
                        fs.tags.append(tag_id)

            # Subscriptions come from ``subscriptions:<RoleType>[:<RoleName>]``
            # columns. The ``subscriptions:<RoleType>`` bare column is a
            # CSV of all emails subscribed in any role of that type —
            # only used as a fallback when no per-role column is set.
            typed_emails: dict[tuple[str, str], set[str]] = {}
            bare_role_type_emails: dict[str, set[str]] = {}
            for col, value in row.items():
                if not (col and col.startswith("subscriptions:")):
                    continue
                parts = col.split(":")
                if len(parts) == 2:
                    role_type = parts[1]
                    bare_role_type_emails.setdefault(role_type, set()).update(_split_csv(value))
                elif len(parts) >= 3:
                    role_type = parts[1]
                    role_name = ":".join(parts[2:])
                    typed_emails.setdefault((role_type, role_name), set()).update(_split_csv(value))

            covered_emails_by_type: dict[str, set[str]] = {}
            for (role_type, role_name), emails in typed_emails.items():
                for email in emails:
                    subscriptions.append(
                        Subscription(
                            leanix_id=f"{fs.leanix_id}:{role_type}:{role_name}:{email}",
                            fact_sheet_id=fs.leanix_id,
                            user_email=email,
                            user_display_name=None,
                            role_name=role_name,
                            role_type=role_type,
                            raw={},
                        )
                    )
                covered_emails_by_type.setdefault(role_type, set()).update(emails)

            for role_type, emails in bare_role_type_emails.items():
                fallback = emails - covered_emails_by_type.get(role_type, set())
                for email in fallback:
                    subscriptions.append(
                        Subscription(
                            leanix_id=f"{fs.leanix_id}:{role_type}::{email}",
                            fact_sheet_id=fs.leanix_id,
                            user_email=email,
                            user_display_name=None,
                            role_name=None,
                            role_type=role_type,
                            raw={},
                        )
                    )

    # Third pass: relation sheets. We also keep track of distinct
    # (rel_type → (from_fs_type, to_fs_type)) tuples so the metamodel
    # synthesizer can emit a MetamodelRelationType for every relation
    # name that ends up unmapped — without this, custom relations
    # involving the 5+ new fact-sheet types (Server, ESGCapability,
    # System, TechPlatform, TechnicalStack on the demo workbook) have
    # nowhere to land.
    rel_type_endpoints: dict[str, tuple[str, str]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in _AUX_SHEETS:
            continue
        sheet = wb[sheet_name]
        first_header = _first_data_header(sheet)
        if first_header is None or "fromRelatedFactSheetDisplayName" not in first_header:
            continue
        for row in _data_rows(sheet, skip_label_row=True):
            try:
                rel = _build_relation(row, display_index, errors)
            except _SkipRowError:
                continue
            if rel is None:
                continue
            relations.append(rel)
            # First-seen endpoint pair wins — relations of the same type
            # always have the same endpoint types in practice.
            if rel.type not in rel_type_endpoints:
                from_t = _str_or_none(row.get("fromRelatedFactSheetType"))
                to_t = _str_or_none(row.get("toRelatedFactSheetType"))
                if from_t and to_t:
                    rel_type_endpoints[rel.type] = (from_t, to_t)

    # Documents + comments (auxiliary sheets — reference fact sheets by display name).
    if "Documents" in wb.sheetnames:
        for row in _data_rows(wb["Documents"], skip_label_row=True):
            fs_id = _resolve_endpoint_id(
                _str_or_none(row.get("factSheet")),
                _str_or_none(row.get("type")),
                display_index,
                errors,
                context=f"document {row.get('id')!r}",
            )
            if not fs_id:
                continue
            documents.append(
                Document(
                    leanix_id=_str_or_none(row.get("id")) or "",
                    fact_sheet_id=fs_id,
                    name=_str_or_none(row.get("name")) or "",
                    url=_str_or_none(row.get("url")),
                    raw={k: v for k, v in row.items() if v not in (None, "")},
                )
            )

    if "Comments" in wb.sheetnames:
        for sheet_index, row in enumerate(_data_rows(wb["Comments"], skip_label_row=True)):
            fs_id = _resolve_endpoint_id(
                _str_or_none(row.get("factSheet")),
                _str_or_none(row.get("type")),
                display_index,
                errors,
                context="comment",
            )
            if not fs_id:
                continue
            body = _str_or_none(row.get("message")) or ""
            if not body:
                continue
            created = _coerce_datetime(row.get("createdAt"))
            author = _str_or_none(row.get("userEmail")) or ""
            comments.append(
                Comment(
                    leanix_id=_synth_comment_id(fs_id, created, author, body, sheet_index),
                    fact_sheet_id=fs_id,
                    author_email=author or None,
                    body=body,
                    created_at=created,
                    raw={k: v for k, v in row.items() if v not in (None, "")},
                )
            )
            # Reply, if present, also lands as a flat top-level comment
            # — the parser does not try to reconstruct LeanIX's UI
            # thread chain (the staging layer drops thread parent links
            # anyway).
            reply_body = _str_or_none(row.get("replyMessage"))
            if reply_body:
                reply_created = _coerce_datetime(row.get("replyCreatedAt"))
                reply_author = _str_or_none(row.get("replierEmail")) or ""
                comments.append(
                    Comment(
                        leanix_id=_synth_comment_id(
                            fs_id, reply_created, reply_author, reply_body, sheet_index, reply=True
                        ),
                        fact_sheet_id=fs_id,
                        author_email=reply_author or None,
                        body=reply_body,
                        created_at=reply_created,
                        raw={},
                    )
                )

    # Resolve hierarchy (childParentRelation → FactSheet.parent_id).
    _apply_child_parent_relations(fact_sheets, relations)

    # Synthesize a metamodel summary so the admin sees every column
    # they're about to import — known LeanIX standard fields are flagged
    # ``is_custom=False`` and skipped by ``stage_metamodel``; everything
    # else surfaces as a new field row.
    metamodel_types = _synthesize_metamodel(seen_columns_per_type, types_options, readme_fields)

    # Every distinct relation type observed in the workbook surfaces as
    # a synthetic MetamodelRelationType. The migration service filters
    # out names already mapped in ``LX_TO_TEA_RELATION`` (those route to
    # an existing Turbo EA edge); what's left becomes a new non-builtin
    # relation type on apply, so custom relations involving the new
    # fact-sheet types (Server, ESGCapability, lxSystem…) actually
    # land in the database.
    metamodel_relation_types = [
        MetamodelRelationType(
            name=rel_type,
            source_type=src,
            target_type=tgt,
            label=rel_type,
            attributes_schema=[],
            is_custom=True,
        )
        for rel_type, (src, tgt) in rel_type_endpoints.items()
        # ``childParentRelation`` is the hierarchy edge — folded into
        # ``Card.parent_id`` by ``_apply_child_parent_relations`` above
        # and must NOT surface as a standalone Turbo EA relation type.
        if rel_type != "childParentRelation"
    ]

    # Distinct users referenced by any subscription (the export carries
    # no separate user sheet).
    users = _distinct_users(subscriptions)

    return LeanixSnapshot(
        version="xlsx",
        fact_sheets=fact_sheets,
        relations=relations,
        subscriptions=subscriptions,
        tags=tag_records,
        documents=documents,
        comments=comments,
        users=users,
        metamodel_types=metamodel_types,
        metamodel_relation_types=metamodel_relation_types,
        parse_errors=errors,
    )


# ---------------------------------------------------------------------------
# Per-row builders
# ---------------------------------------------------------------------------


class _SkipRowError(Exception):
    """Sentinel used by row builders to bail without aborting the sheet."""


def _build_fact_sheet(row: dict[str, Any], errors: list[str]) -> FactSheet:
    leanix_id = _str_or_none(row.get("id"))
    fs_type = _str_or_none(row.get("type"))
    if not leanix_id or not fs_type:
        raise _SkipRowError
    name = _str_or_none(row.get("name")) or _str_or_none(row.get("displayName")) or leanix_id
    display_name = _str_or_none(row.get("displayName")) or name

    lifecycle: dict[str, str] = {}
    for col, value in row.items():
        if not (col and col.startswith("lifecycle:")):
            continue
        phase = col[len("lifecycle:") :]
        dt = _coerce_datetime(value)
        if dt is not None:
            lifecycle[phase] = dt.date().isoformat()

    custom_fields: dict[str, Any] = {}
    for col, value in row.items():
        if not col or col in _FS_CORE_COLS:
            continue
        if col.startswith(("lifecycle:", "tags:", "subscriptions:")):
            continue
        # Skip empty cells — they carry no signal.
        if value in (None, ""):
            continue
        custom_fields[col] = _jsonify(value)

    return FactSheet(
        leanix_id=leanix_id,
        type=fs_type,
        name=name,
        display_name=display_name,
        category=_str_or_none(row.get("category")),
        description=_str_or_none(row.get("description")),
        lifecycle=lifecycle,
        tags=[],
        parent_id=None,  # set by _apply_child_parent_relations
        custom_fields=custom_fields,
        quality_seal=_str_or_none(row.get("qualitySeal")),
        completion=_float_or_none(row.get("completion")),
        status=_str_or_none(row.get("status")),
        raw={k: v for k, v in row.items() if v not in (None, "")},
    )


def _build_relation(
    row: dict[str, Any],
    display_index: dict[tuple[str, str], str],
    errors: list[str],
) -> Relation | None:
    rel_id = _str_or_none(row.get("id"))
    rel_type = _str_or_none(row.get("type"))
    if not rel_type:
        raise _SkipRowError
    from_name = _str_or_none(row.get("fromRelatedFactSheetDisplayName"))
    from_type = _str_or_none(row.get("fromRelatedFactSheetType"))
    to_name = _str_or_none(row.get("toRelatedFactSheetDisplayName"))
    to_type = _str_or_none(row.get("toRelatedFactSheetType"))
    if not (from_name and from_type and to_name and to_type):
        raise _SkipRowError

    # LeanIX exports two flavours of successor edges in the same workbook:
    # type-prefixed (``applicationSuccessorRelation``) and the generic
    # ``successorRelation`` — the latter relies on the row's
    # ``fromRelatedFactSheetType`` to disambiguate which TEA edge it maps
    # to. Rewrite the generic form to the prefixed form at parse time so
    # the static ``LX_TO_TEA_RELATION`` map and the ``LX_FLIP_DIRECTION``
    # set catch them just like the prefixed rows. Same treatment for the
    # GraphQL ``relSuccessor`` equivalent. Mixed-type successor rows
    # (extremely rare in LeanIX) are left untouched so they surface as
    # admin-reviewable conflicts.
    rel_type = _specialise_generic_successor(rel_type, from_type, to_type)

    src = _resolve_endpoint_id(from_name, from_type, display_index, errors, context=rel_type)
    tgt = _resolve_endpoint_id(to_name, to_type, display_index, errors, context=rel_type)
    if not (src and tgt):
        # Dangling endpoint — record as parse error but don't emit the
        # Relation (the staging layer would mark it conflict anyway).
        errors.append(
            f"relation {rel_type}: dangling endpoint "
            f"from=({from_name!r},{from_type}) → "
            f"to=({to_name!r},{to_type})"
        )
        return None

    attributes: dict[str, Any] = {}
    for col, value in row.items():
        if not col or col in _REL_CORE_COLS:
            continue
        if value in (None, ""):
            continue
        attributes[col] = _jsonify(value)

    return Relation(
        leanix_id=rel_id or f"{rel_type}:{src}:{tgt}",
        type=rel_type,
        source_id=src,
        target_id=tgt,
        attributes=attributes,
        raw={k: v for k, v in row.items() if v not in (None, "")},
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _data_rows(sheet: Any, *, skip_label_row: bool) -> list[dict[str, Any]]:
    """Iterate a sheet row-by-row as ``{header: value}`` dicts.

    Row 1 holds the LeanIX field keys; row 2 holds the human-readable
    "label" header (e.g. ``ID``, ``Type``). Pass
    ``skip_label_row=True`` to skip row 2.
    """
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    headers_norm: list[str | None] = [(str(h).strip() if h is not None else None) for h in headers]
    if skip_label_row:
        next(rows, None)
    out: list[dict[str, Any]] = []
    for raw_row in rows:
        if not any(v not in (None, "") for v in raw_row):
            continue
        out.append({h: v for h, v in zip(headers_norm, raw_row) if h is not None})
    return out


def _first_data_header(sheet: Any) -> set[str] | None:
    """Peek the first row of a sheet and return its header set, or None if empty."""
    rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    headers = next(iter(rows), None)
    if not headers:
        return None
    return {str(h).strip() for h in headers if h is not None}


def _str_or_none(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _synth_comment_id(
    fs_id: str,
    created: datetime | None,
    author: str,
    body: str,
    sheet_index: int,
    *,
    reply: bool = False,
) -> str:
    """Build a stable unique id for an xlsx comment row.

    The LeanIX xlsx export carries no id column on Comments, so the
    parser synthesises one. Requirements:

    1. **Deterministic** — the same row produces the same id across
       re-uploads so the identity map stays meaningful.
    2. **Collision-free** within a single migration — two comments
       posted on the same fact sheet at the same `createdAt` second by
       distinct authors (or even two identical lines posted twice by
       the same author) must not share an id.

    Strategy: hash (fs_id, ts, author, body, sheet_index, reply-flag)
    with MD5 truncated to 12 hex chars (~48 bits, 1e-8 collision odds
    over 5,000 comments) and prefix with the fact-sheet uuid + ts so
    debug output stays human-readable. The ``sheet_index`` is the
    Comments-sheet row position — re-running the same export keeps the
    same order, so ids remain stable across re-uploads.
    """
    ts = created.isoformat() if created else "na"
    digest_input = f"{fs_id}|{ts}|{author}|{sheet_index}|{int(reply)}|{body}"
    digest = hashlib.md5(digest_input.encode("utf-8")).hexdigest()[:12]
    flag = "reply" if reply else "msg"
    return f"{fs_id}:{ts}:{flag}:{digest}"


def _float_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _coerce_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if value in (None, ""):
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return None


def _split_csv(value: Any) -> list[str]:
    """Split a comma-separated LeanIX cell into a list of trimmed values.

    LeanIX never quotes its CSV cells in xlsx exports — a literal comma
    inside a tag name will be misread, but that's also how LeanIX's own
    upload pipeline reads them back, so we mirror the behaviour.
    """
    s = _str_or_none(value)
    if not s:
        return []
    return [chunk.strip() for chunk in s.split(",") if chunk.strip()]


def _norm_name(value: Any) -> str:
    s = _str_or_none(value)
    return s or ""


def _resolve_endpoint_id(
    display_name: str | None,
    fs_type: str | None,
    display_index: dict[tuple[str, str], str],
    errors: list[str],
    *,
    context: str = "",
) -> str | None:
    if not (display_name and fs_type):
        return None
    leanix_id = display_index.get((display_name.strip(), fs_type.strip()))
    if leanix_id is None:
        errors.append(f"{context}: unresolved endpoint ({display_name!r}, {fs_type!r})")
    return leanix_id


# LeanIX fact-sheet type names → camelCased prefix the type-specific
# successor edge uses (``Application`` → ``application`` →
# ``applicationSuccessorRelation``). The keys cover both LeanIX's
# canonical names and the tenant aliases ``Project`` (folded into
# Initiative downstream) and ``TechPlatform`` (folded into Platform);
# the camelCased ``itComponent`` is preserved as-is.
_SUCCESSOR_TYPE_PREFIX: dict[str, str] = {
    "Application": "application",
    "ITComponent": "itComponent",
    "Interface": "interface",
    "Initiative": "initiative",
    "Project": "initiative",
    "Platform": "platform",
    "TechPlatform": "platform",
    "BusinessProcess": "process",
    "Process": "process",
    "DataObject": "dataObject",
}


def _specialise_generic_successor(rel_type: str, from_type: str | None, to_type: str | None) -> str:
    """Rewrite LeanIX's generic ``successorRelation`` / ``relSuccessor``
    rows into the type-specific form (``applicationSuccessorRelation``
    etc.) so the downstream mapping + direction-flip layers can route
    them without a special case. Returns ``rel_type`` unchanged when the
    row isn't a generic successor edge or when the endpoints have
    different FS types (an invalid LeanIX shape we leave alone for the
    admin to inspect)."""
    if rel_type not in {"successorRelation", "relSuccessor"}:
        return rel_type
    if not from_type or from_type != to_type:
        return rel_type
    prefix = _SUCCESSOR_TYPE_PREFIX.get(from_type)
    if not prefix:
        return rel_type
    return f"{prefix}SuccessorRelation"


def _apply_child_parent_relations(
    fact_sheets: list[FactSheet],
    relations: list[Relation],
) -> None:
    """Translate ``childParentRelation`` rows into ``FactSheet.parent_id``.

    Direction in the LeanIX export: ``from = child``, ``to = parent``.
    We drop the relation from the list after applying so the staging
    layer doesn't try to re-create it as a Turbo EA relation (TEA uses
    ``Card.parent_id`` exclusively for hierarchy).
    """
    by_id = {fs.leanix_id: fs for fs in fact_sheets}
    remaining: list[Relation] = []
    for rel in relations:
        if rel.type == "childParentRelation":
            child = by_id.get(rel.source_id)
            if child is not None:
                child.parent_id = rel.target_id
            continue
        remaining.append(rel)
    relations[:] = remaining


def _synthesize_metamodel(
    seen_columns: dict[str, dict[str, list[Any]]],
    types_options: dict[tuple[str, str], list[str]],
    readme_fields: dict[tuple[str, str], dict[str, Any]],
) -> list[MetamodelType]:
    """Build a fake ``MetamodelType[]`` from observed column headers.

    Resolution order for each (fs_type, column) pair:

    1. ``readme_fields[(fs_type, col)]`` — the per-type entry from the
       ReadMe sheet (authoritative type + full enum).
    2. ``readme_fields[("*", col)]`` — the "All Fact Sheet types"
       entry from ReadMe, for fields like ``externalId`` or
       ``signavioGlossaryItemId`` that apply globally.
    3. ``types_options[(fs_type, col)]`` — fallback enum scraped from
       the in-data ``Types`` sheet (only contains values actually
       observed on cards, so may be a strict subset of the real enum).
    4. Default: ``data_type="STRING"`` with no options.

    Standard LeanIX columns (``name``, ``description``, lifecycle
    phases, tag/subscription columns…) are filtered out so the preview
    only highlights what is actually new.
    """
    out: list[MetamodelType] = []
    for fs_type, cols in seen_columns.items():
        fields: list[MetamodelField] = []
        for col in sorted(cols):
            if col in _FS_CORE_COLS:
                continue
            if col.startswith(("lifecycle:", "tags:", "subscriptions:")):
                continue
            readme = readme_fields.get((fs_type, col)) or readme_fields.get(("*", col))
            if readme is not None:
                data_type = readme["data_type"]
                options_list = list(readme.get("options") or [])
            else:
                options_list = [
                    {"key": v, "label": v} for v in types_options.get((fs_type, col), [])
                ]
                data_type = "SINGLE_SELECT" if options_list else "STRING"
            fields.append(
                MetamodelField(
                    type_name=fs_type,
                    key=col,
                    label=(readme or {}).get("label") or col,
                    data_type=data_type,
                    options=options_list,
                    translations={},
                    is_custom=True,
                )
            )
        out.append(
            MetamodelType(
                name=fs_type,
                is_custom=True,  # surfaced as ``metamodel_type`` for the admin
                fields=fields,
                subtypes=[],
            )
        )
    return out


# ---------------------------------------------------------------------------
# ReadMe reference parser
# ---------------------------------------------------------------------------


# ReadMe ``Type`` column → LeanIX ``data_type`` value the migration
# service can route through :data:`LX_DATATYPE_TO_TEA_TYPE`. We keep
# the LeanIX-flavoured names rather than already collapsing to
# ``text``/``number``/… here, so the migration service stays the
# single point that maps LX → TEA.
_README_TYPE_MAP: dict[str, str] = {
    "string": "STRING",
    "string (uuid)": "STRING",
    "string list": "MULTIPLE_SELECT",
    "integer": "INTEGER",
    "percent": "DOUBLE",
    "double": "DOUBLE",
    "money": "MONEY",
    "datetime": "DATETIME",
    "date": "DATE",
    "boolean": "BOOLEAN",
    "url": "URL",
    "email address": "EMAIL",
}


def _parse_readme(wb: Any) -> dict[tuple[str, str], dict[str, Any]]:
    """Build a ``(fs_type_or_'*', col_name) → {data_type, options, …}`` index from the ReadMe.

    The ReadMe sheet groups field definitions under section markers
    that look like:

    - ``All Fact Sheet types`` — applies to every type (scope ``*``)
    - ``Fact Sheet type Application`` — applies only to Applications

    Every following data row carries
    ``(column, lx_type, mandatory, read_only, remarks)``. Enum
    constraints are encoded in the remarks as
    ``Possible values: [one of] X, Y, Z.`` which we lift back into a
    structured options list.
    """
    out: dict[tuple[str, str], dict[str, Any]] = {}
    if "ReadMe" not in wb.sheetnames:
        return out
    ws = wb["ReadMe"]
    scope: str | None = None
    seen_header = False
    for row in ws.iter_rows(values_only=True):
        col_a = _str_or_none(row[0]) if len(row) > 0 else None
        col_b = _str_or_none(row[1]) if len(row) > 1 else None
        col_e = _str_or_none(row[4]) if len(row) > 4 else None
        if not col_a:
            continue
        # Header row marks the start of structured definitions.
        if col_a == "Column header" and col_b == "Type":
            seen_header = True
            continue
        if not seen_header:
            continue
        # Section marker rows have ONLY column A set.
        if col_a and not any(_str_or_none(c) for c in row[1:]):
            if col_a == "All Fact Sheet types":
                scope = "*"
            elif col_a.startswith("Fact Sheet type "):
                scope = col_a[len("Fact Sheet type ") :].strip()
            elif col_a == "Relations to Sheets Mapping":
                break  # tail section unrelated to fields
            else:
                scope = None  # unknown section — stop emitting until a known one
            continue
        if scope is None or not col_b:
            continue
        # Field-definition row.
        lx_type = _README_TYPE_MAP.get(col_b.strip().lower(), "STRING")
        options = _extract_possible_values(col_e or "")
        # If the remarks list options but the type is ``String`` (the
        # default LeanIX rendering for any enum), promote to
        # ``SINGLE_SELECT`` so the TEA field type is a dropdown.
        if options and lx_type == "STRING":
            lx_type = "SINGLE_SELECT"
        out[(scope, col_a)] = {
            "data_type": lx_type,
            "options": [{"key": v, "label": v} for v in options],
            "label": col_a,
            "mandatory": _str_or_none(row[2] if len(row) > 2 else None) == "mandatory",
        }
    return out


def _extract_possible_values(remarks: str) -> list[str]:
    """Lift ``Possible values: [one of ]X, Y, Z.`` out of a remarks blurb.

    Returns an empty list when no such constraint is present.
    """
    marker = "Possible values:"
    if marker not in remarks:
        return []
    tail = remarks.split(marker, 1)[1].strip()
    # Drop the trailing period (and any text past it — sometimes there
    # is follow-up prose after the enum).
    tail = tail.split(".", 1)[0]
    # Strip optional leading "one of".
    if tail.lower().startswith("one of"):
        tail = tail[len("one of") :].strip()
    return [v.strip() for v in tail.split(",") if v.strip()]


def _distinct_users(subscriptions: list[Subscription]) -> list[UserRef]:
    seen: dict[str, UserRef] = {}
    for sub in subscriptions:
        email = (sub.user_email or "").strip().lower()
        if not email or email in seen:
            continue
        seen[email] = UserRef(
            leanix_id=email,
            email=email,
            display_name=sub.user_display_name or email,
        )
    return list(seen.values())


def _jsonify(value: Any) -> Any:
    """Convert a cell value into a JSON-serialisable scalar."""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value
