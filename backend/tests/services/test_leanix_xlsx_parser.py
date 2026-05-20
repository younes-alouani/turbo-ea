"""Unit tests for the LeanIX xlsx-export parser.

The tests build small synthetic workbooks in-memory (never real
customer data) and assert that the dynamic columns LeanIX's export
emits — ``lifecycle:<phase>``, ``tags:<group>``,
``subscriptions:<roleType>[:<roleName>]`` — translate cleanly into the
shared :class:`LeanixSnapshot` dataclass.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook  # type: ignore[import-untyped]

from app.services.leanix_xlsx_parser import is_xlsx_payload, parse_xlsx, parse_xlsx_path


def _write_sheet(wb: Workbook, name: str, rows: list[list]) -> None:
    """Helper — append rows to a sheet, replacing the default sheet on first call."""
    if name in wb.sheetnames:
        ws = wb[name]
    elif wb.sheetnames == ["Sheet"]:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)
    for row in rows:
        ws.append(row)


def _to_stream(wb: Workbook) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _minimal_workbook() -> Workbook:
    """Build a workbook with one Application, one BC, and a child→parent link."""
    wb = Workbook()
    # Application sheet — note the dynamic ``lifecycle:``, ``tags:``,
    # ``subscriptions:`` columns plus a tenant-custom column.
    _write_sheet(
        wb,
        "Application",
        [
            [
                "id",
                "type",
                "name",
                "displayName",
                "status",
                "description",
                "completion",
                "qualitySeal",
                "lifecycle:plan",
                "lifecycle:active",
                "tags:Region",
                "tags:Other tags",
                "subscriptions:RESPONSIBLE:Application Owner",
                "subscriptions:OBSERVER",
                "customRenewalDate",
            ],
            [
                "ID",
                "Type",
                "Name",
                "Display Name",
                "Status",
                "Description",
                "Completion",
                "QualitySeal",
                "Lifecycle: plan",
                "Lifecycle: active",
                "Tags: Region",
                "Tags: Other tags",
                "Subscriptions: RESPONSIBLE: Application Owner",
                "Subscriptions OBSERVER",
                "customRenewalDate",
            ],
            [
                "fs-app-1",
                "Application",
                "Salesforce",
                "Salesforce CRM",
                "ACTIVE",
                "CRM",
                0.85,
                "GREEN",
                datetime(2019, 1, 1),
                datetime(2020, 1, 1),
                "EMEA, APAC",
                "Pilot",
                "owner@example.com",
                "watcher@example.com, second@example.com",
                "2027-06-30",
            ],
        ],
    )
    # BusinessCapability sheet
    _write_sheet(
        wb,
        "BusinessCapability",
        [
            ["id", "type", "name", "displayName", "status", "description"],
            ["ID", "Type", "Name", "Display Name", "Status", "Description"],
            ["fs-bc-parent", "BusinessCapability", "Sales", "Sales", "ACTIVE", ""],
            ["fs-bc-child", "BusinessCapability", "Lead Mgmt", "Lead Mgmt", "ACTIVE", ""],
        ],
    )
    # Hierarchy edge (child → parent)
    _write_sheet(
        wb,
        "childParentRelation",
        [
            [
                "id",
                "type",
                "fromRelatedFactSheetDisplayName",
                "fromRelatedFactSheetType",
                "toRelatedFactSheetDisplayName",
                "toRelatedFactSheetType",
                "status",
            ],
            ["ID", "Type", "From", "From Type", "To", "To Type", "Status"],
            [
                "rel-1",
                "childParentRelation",
                "Lead Mgmt",
                "BusinessCapability",
                "Sales",
                "BusinessCapability",
                "ACTIVE",
            ],
        ],
    )
    # An ordinary relation sheet
    _write_sheet(
        wb,
        "applicationBusinessCapability",  # name is truncated by Excel — fine
        [
            [
                "id",
                "type",
                "fromRelatedFactSheetDisplayName",
                "fromRelatedFactSheetType",
                "toRelatedFactSheetDisplayName",
                "toRelatedFactSheetType",
                "status",
                "technicalSuitability",
            ],
            ["ID", "Type", "From", "From Type", "To", "To Type", "Status", "TS"],
            [
                "rel-2",
                "applicationBusinessCapabilityRelation",
                "Salesforce CRM",
                "Application",
                "Lead Mgmt",
                "BusinessCapability",
                "ACTIVE",
                "appropriate",
            ],
        ],
    )
    # Tag metadata sheets
    _write_sheet(
        wb,
        "TagGroups",
        [
            [
                "id",
                "name",
                "shortName",
                "description",
                "mode",
                "numberOfTags",
                "restrictToFactSheetTypes",
            ],
            ["ID", "Name", "Short", "Desc", "Mode", "Count", "Allowed"],
            ["tg-1", "Region", "", "", "MULTIPLE", 3, ""],
            ["tg-2", "Other tags", "", "", "SINGLE", 1, ""],
        ],
    )
    _write_sheet(
        wb,
        "Tags",
        [
            [
                "id",
                "tagGroupId",
                "name",
                "description",
                "backgroundColor",
                "status",
                "factSheetNumber",
            ],
            ["ID", "Group", "Name", "Desc", "Color", "Status", "FS#"],
            ["tag-emea", "Region", "EMEA", "", "#1194e0", "ACTIVE", 10],
            ["tag-apac", "Region", "APAC", "", "#1194e0", "ACTIVE", 5],
            ["tag-pilot", "Other tags", "Pilot", "", "#999999", "ACTIVE", 1],
        ],
    )
    # Documents + comments — both reference FS by display name
    _write_sheet(
        wb,
        "Documents",
        [
            [
                "id",
                "factSheet",
                "name",
                "description",
                "url",
                "createdAt",
                "filename",
                "mediaType",
                "documentType",
                "metadata",
                "refId",
                "type",
            ],
            [
                "ID",
                "FS",
                "Name",
                "Desc",
                "URL",
                "At",
                "File",
                "Media",
                "DocType",
                "Meta",
                "Ref",
                "Type",
            ],
            [
                "doc-1",
                "Salesforce CRM",
                "Runbook",
                "",
                "https://wiki/runbook",
                datetime(2024, 6, 1),
                "",
                "",
                "documentation",
                "",
                "",
                "Application",
            ],
        ],
    )
    _write_sheet(
        wb,
        "Comments",
        [
            [
                "factSheet",
                "createdAt",
                "message",
                "userEmail",
                "status",
                "replierEmail",
                "replyCreatedAt",
                "replyMessage",
                "type",
            ],
            ["FS", "At", "Msg", "By", "Status", "ReplyBy", "ReplyAt", "ReplyMsg", "Type"],
            [
                "Salesforce CRM",
                datetime(2024, 7, 1, 12, 0, 0),
                "First comment",
                "a@example.com",
                "ACTIVE",
                "b@example.com",
                datetime(2024, 7, 1, 12, 5, 0),
                "First reply",
                "Application",
            ],
        ],
    )
    return wb


def test_xlsx_detection_by_magic_bytes() -> None:
    buf = _to_stream(_minimal_workbook())
    head = buf.read(8)
    assert is_xlsx_payload(head)
    assert not is_xlsx_payload(b'{"version":"2.0"}')
    assert not is_xlsx_payload(b"\x1f\x8b\x08")  # gzip magic


def test_parses_factsheets_with_dynamic_columns() -> None:
    snap = parse_xlsx(_to_stream(_minimal_workbook()))
    assert snap.version == "xlsx"
    assert len(snap.fact_sheets) == 3
    apps = [fs for fs in snap.fact_sheets if fs.type == "Application"]
    assert len(apps) == 1
    fs = apps[0]
    assert fs.leanix_id == "fs-app-1"
    assert fs.name == "Salesforce"
    assert fs.display_name == "Salesforce CRM"
    assert fs.quality_seal == "GREEN"
    assert fs.completion == 0.85
    assert fs.lifecycle == {"plan": "2019-01-01", "active": "2020-01-01"}
    # Tenant-custom column survives untouched, datetimes are
    # JSON-serialisable strings.
    assert fs.custom_fields["customRenewalDate"] == "2027-06-30"
    # ``lifecycle:`` / ``tags:`` / ``subscriptions:`` columns must not
    # leak into custom_fields.
    assert not any(
        k.startswith(("lifecycle:", "tags:", "subscriptions:")) for k in fs.custom_fields
    )


def test_tags_resolve_via_group_name_lookup() -> None:
    snap = parse_xlsx(_to_stream(_minimal_workbook()))
    app = next(fs for fs in snap.fact_sheets if fs.type == "Application")
    # ``tags:Region`` cell was "EMEA, APAC" — expect both ids resolved
    # via the Tags sheet, plus the "Pilot" tag from the second group.
    assert sorted(app.tags) == sorted(["tag-emea", "tag-apac", "tag-pilot"])
    # Tag metadata is exposed at the snapshot level too.
    assert {t.leanix_id for t in snap.tags} == {"tag-emea", "tag-apac", "tag-pilot"}
    region = next(t for t in snap.tags if t.leanix_id == "tag-emea")
    assert region.group_name == "Region"
    assert region.group_mode == "MULTIPLE"
    assert region.color == "#1194e0"


def test_subscriptions_split_per_role_and_dedupe_bare_aggregates() -> None:
    snap = parse_xlsx(_to_stream(_minimal_workbook()))
    # One typed RESPONSIBLE subscription + two bare OBSERVER subscriptions
    # (no per-role observer column, so role_name is None).
    by_email = {(s.user_email, s.role_type, s.role_name) for s in snap.subscriptions}
    assert ("owner@example.com", "RESPONSIBLE", "Application Owner") in by_email
    assert ("watcher@example.com", "OBSERVER", None) in by_email
    assert ("second@example.com", "OBSERVER", None) in by_email
    # Distinct users are deduped across subscriptions.
    assert {u.email for u in snap.users} == {
        "owner@example.com",
        "watcher@example.com",
        "second@example.com",
    }


def test_child_parent_relation_becomes_parent_id() -> None:
    snap = parse_xlsx(_to_stream(_minimal_workbook()))
    child = next(fs for fs in snap.fact_sheets if fs.leanix_id == "fs-bc-child")
    assert child.parent_id == "fs-bc-parent"
    # The relation itself must NOT be emitted — TEA uses Card.parent_id
    # and the staging layer should not see a duplicate edge.
    assert not any(r.type == "childParentRelation" for r in snap.relations)
    # And it must NOT surface as a new metamodel relation type either.
    assert not any(rt.name == "childParentRelation" for rt in snap.metamodel_relation_types)


def test_readme_sheet_lifts_full_enum_options() -> None:
    """The ReadMe sheet is LeanIX's authoritative field reference.

    The data-only Types sheet can underrepresent enums (only the values
    actually used by cards appear). The ReadMe carries the *complete*
    constraint via ``Possible values: one of X, Y, Z.`` — the parser
    must lift those into the synthesised metamodel even when the data
    references only a subset.
    """
    wb = Workbook()
    _write_sheet(
        wb,
        "Application",
        [
            ["id", "type", "name", "displayName", "status", "currentMaturity"],
            ["ID", "Type", "Name", "Display Name", "Status", "Maturity"],
            ["fs-app-1", "Application", "App1", "App1", "ACTIVE", "adHoc"],
        ],
    )
    # Minimal ReadMe sheet documenting currentMaturity with the full
    # 5-value enum. Real LeanIX exports use exactly this shape.
    _write_sheet(
        wb,
        "ReadMe",
        [
            ["LeanIX full export note", None, None, None, None],
            [None, None, None, None, None],
            ["Column header", "Type", "Mandatory", "Read only", "Remarks"],
            ["Fact Sheet type Application", None, None, None, None],
            [
                "currentMaturity",
                "String",
                "",
                "",
                "Possible values: one of adHoc, repeatable, defined, managed, optimized.",
            ],
        ],
    )
    snap = parse_xlsx(_to_stream(wb))
    app = next(mt for mt in snap.metamodel_types if mt.name == "Application")
    field = next(f for f in app.fields if f.key == "currentMaturity")
    # Enum upgraded to SINGLE_SELECT and carries every value from
    # ReadMe, not just the one observed in the data ("adHoc").
    assert field.data_type == "SINGLE_SELECT"
    assert [o["key"] for o in field.options] == [
        "adHoc",
        "repeatable",
        "defined",
        "managed",
        "optimized",
    ]


def test_synthesizes_metamodel_relation_type_for_every_observed_edge() -> None:
    """Every distinct relation type observed in the workbook must surface
    as a ``MetamodelRelationType`` so the migration service can create a
    new Turbo EA relation type for any name it doesn't already map."""
    snap = parse_xlsx(_to_stream(_minimal_workbook()))
    rel_type_names = {rt.name for rt in snap.metamodel_relation_types}
    assert "applicationBusinessCapabilityRelation" in rel_type_names
    rt = next(
        rt
        for rt in snap.metamodel_relation_types
        if rt.name == "applicationBusinessCapabilityRelation"
    )
    # Endpoint LX-type names come straight from the relation row.
    assert rt.source_type == "Application"
    assert rt.target_type == "BusinessCapability"
    assert rt.is_custom is True


def test_relation_endpoints_resolve_by_display_name_and_type() -> None:
    snap = parse_xlsx(_to_stream(_minimal_workbook()))
    rels = [r for r in snap.relations if r.type == "applicationBusinessCapabilityRelation"]
    assert len(rels) == 1
    rel = rels[0]
    assert rel.source_id == "fs-app-1"
    assert rel.target_id == "fs-bc-child"
    # Attributes outside the standard relation columns flow through.
    assert rel.attributes == {"technicalSuitability": "appropriate"}


def test_documents_and_comments_resolve_factsheet_by_display_name() -> None:
    snap = parse_xlsx(_to_stream(_minimal_workbook()))
    assert len(snap.documents) == 1
    doc = snap.documents[0]
    assert doc.fact_sheet_id == "fs-app-1"
    assert doc.url == "https://wiki/runbook"

    # Comment + reply both land as flat comments on the same FS.
    bodies = sorted(c.body for c in snap.comments)
    assert bodies == ["First comment", "First reply"]
    assert all(c.fact_sheet_id == "fs-app-1" for c in snap.comments)


def test_synthesized_metamodel_surfaces_tenant_custom_columns() -> None:
    snap = parse_xlsx(_to_stream(_minimal_workbook()))
    # The Application type carries one tenant-custom column we put in
    # the sheet (``customRenewalDate``). That should appear in the
    # synthesized metamodel so :func:`stage_metamodel` can emit a new
    # field row for it.
    app_type = next(t for t in snap.metamodel_types if t.name == "Application")
    field_keys = {f.key for f in app_type.fields}
    assert "customRenewalDate" in field_keys
    # Core columns must NOT leak into the synthesized custom-field list.
    assert "name" not in field_keys
    assert "description" not in field_keys


def test_unknown_factsheet_type_passes_through_for_admin_review() -> None:
    """A custom (tenant-defined) FS type must parse without dropping rows.

    The migration_service then surfaces it as a ``metamodel_type``
    staged record so the admin can map it to an existing TEA card type
    or create a new one.
    """
    wb = Workbook()
    _write_sheet(
        wb,
        "ESGCapability",
        [
            ["id", "type", "name", "displayName", "status", "description"],
            ["ID", "Type", "Name", "Display Name", "Status", "Description"],
            ["fs-esg-1", "ESGCapability", "Carbon Reporting", "Carbon Reporting", "ACTIVE", ""],
        ],
    )
    snap = parse_xlsx(_to_stream(wb))
    assert len(snap.fact_sheets) == 1
    assert snap.fact_sheets[0].type == "ESGCapability"
    assert any(t.name == "ESGCapability" for t in snap.metamodel_types)


def test_parse_xlsx_path_loads_workbook_regardless_of_extension(tmp_path) -> None:
    """``parse_xlsx_path`` must load xlsx files even when the path has
    no ``.xlsx`` extension — uploaded snapshots are stored under a
    neutral ``.bin`` suffix so openpyxl's extension check doesn't kick
    in, and content-based loading is the only path that works for both
    ``upload.xlsx`` and stored snapshots."""
    wb_path = tmp_path / "export.bin"
    wb = _minimal_workbook()
    wb.save(str(wb_path))
    snap = parse_xlsx_path(str(wb_path))
    assert snap.version == "xlsx"
    assert len(snap.fact_sheets) == 3
